#!/usr/bin/env python3
"""Excel spreadsheet creator using openpyxl."""

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    from echo_agent.dependencies import require
    require("skill.excel-author")
except ImportError:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("Install: pip install openpyxl")


def from_csv(csv_file, output="output.xlsx", sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    with open(csv_file, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            ws.append(row)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    wb.save(output)
    print(f"Created: {output}")


def from_json(json_file, output="output.xlsx"):
    data = json.loads(Path(json_file).read_text())
    if not data:
        sys.exit("Empty JSON data")
    wb = Workbook()
    ws = wb.active
    headers = list(data[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for row in data:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    wb.save(output)
    print(f"Created: {output} ({len(data)} rows)")


def quick_table(title, headers, rows, output="output.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    wb.save(output)
    print(f"Created: {output}")


def main():
    parser = argparse.ArgumentParser(description="Excel creator")
    sub = parser.add_subparsers(dest="cmd")
    p = sub.add_parser("from-csv")
    p.add_argument("input")
    p.add_argument("--output", "-o", default="output.xlsx")
    p = sub.add_parser("from-json")
    p.add_argument("input")
    p.add_argument("--output", "-o", default="output.xlsx")
    p = sub.add_parser("quick")
    p.add_argument("title")
    p.add_argument("--headers", nargs="+", required=True)
    p.add_argument("--output", "-o", default="output.xlsx")
    args = parser.parse_args()

    if args.cmd == "from-csv":
        from_csv(args.input, args.output)
    elif args.cmd == "from-json":
        from_json(args.input, args.output)
    elif args.cmd == "quick":
        quick_table(args.title, args.headers, [], args.output)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
